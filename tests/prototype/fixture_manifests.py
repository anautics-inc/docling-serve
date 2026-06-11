from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


def manifest_for_file(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return docx_manifest(path)
    if suffix == ".xlsx":
        return xlsx_manifest(path)
    if suffix == ".pdf":
        return pdf_manifest(path)
    raise ValueError(f"Unsupported fixture type: {path}")


def docx_manifest(path: Path) -> dict[str, Any]:
    document = Document(path)
    units = []
    current: dict[str, Any] | None = None
    for paragraph in document.paragraphs:
        text = paragraph.text.strip()
        if not text:
            continue
        style = paragraph.style.name if paragraph.style else ""
        is_heading = style.lower().startswith("heading") or current is None
        if is_heading:
            if current is not None:
                units.append(current)
            current = unit(path, len(units), "section", text)
            continue
        assert current is not None
        current["elements"].append(text_element(current["unitId"], len(current["elements"]), text))
    if current is not None:
        units.append(current)
    return manifest(path, "docx", units)


def xlsx_manifest(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    units = []
    for index, worksheet in enumerate(workbook.worksheets):
        values: list[str] = []
        table_rows: list[list[Any]] = []
        for row in worksheet.iter_rows(values_only=True):
            cleaned = [cell for cell in row if cell is not None]
            if not cleaned:
                continue
            table_rows.append(list(cleaned))
            values.append(" | ".join(str(cell) for cell in cleaned))
        item = unit(path, index, "sheet", worksheet.title)
        item["elements"].append(
            {
                "elementId": f"{item['unitId']}-table-001",
                "type": "table",
                "kind": "table",
                "text": {"plain": "\n".join(values), "paragraphs": [], "runs": []},
                "rows": table_rows,
                "bbox": {},
                "source": {"sheetName": worksheet.title},
                "quality": {"reviewRequired": False},
            }
        )
        units.append(item)
    return manifest(path, "xlsx", units)


def pdf_manifest(path: Path) -> dict[str, Any]:
    reader = PdfReader(path)
    units = []
    for index, page in enumerate(reader.pages):
        text = (page.extract_text() or "").strip()
        item = unit(path, index, "page", first_line(text) or f"Page {index + 1}")
        item["elements"].append(text_element(item["unitId"], 0, text))
        units.append(item)
    return manifest(path, "pdf", units)


def manifest(path: Path, file_type: str, units: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "artifactKind": f"captify.{file_type}.fixtureExtraction.v1",
        "source": {
            "path": str(path),
            "originalFileName": path.name,
            "sha256": sha256_file(path),
            "sizeBytes": path.stat().st_size,
            "rendererUsed": False,
            "conversionUsed": False,
            "watermarkRisk": False,
        },
        "stats": {
            "unitCount": len(units),
            "fileType": file_type,
        },
        "units": units,
        "assets": [],
    }


def unit(path: Path, index: int, unit_type: str, title: str) -> dict[str, Any]:
    unit_id = f"{path.suffix.lower().lstrip('.')}-{index + 1:03d}"
    return {
        "unitId": unit_id,
        "unitType": unit_type,
        "index": index,
        "title": title,
        "speakerNotes": {"raw": "", "cleaned": ""},
        "elements": [],
        "background": {},
        "instructionalMetadata": {
            "bloom": {
                "taxonomy": "Bloom",
                "primaryLevel": "Understand",
                "confidence": "low",
                "method": "fixture_extraction",
                "status": "needs_llm_review",
                "evidence": [],
            }
        },
    }


def text_element(unit_id: str, index: int, text: str) -> dict[str, Any]:
    return {
        "elementId": f"{unit_id}-text-{index + 1:03d}",
        "type": "text",
        "kind": "text",
        "text": {
            "plain": text,
            "paragraphs": [{"runs": [{"text": text, "font": {}, "color": None}]}],
            "runs": [{"text": text, "font": {}, "color": None}],
        },
        "bbox": {},
        "source": {},
        "quality": {"reviewRequired": False},
    }


def first_line(text: str) -> str:
    for line in text.splitlines():
        if line.strip():
            return line.strip()
    return ""


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()

